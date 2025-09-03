
#!/usr/bin/env python3
"""
GCP Audit Evidence Demo - Educational Version
============================================
Demonstrates automated transformation of GCP security data to audit-ready Excel
Includes CISA ScubaGoggles integration with professional formatting
"""

import json
import pandas as pd # pyright: ignore[reportMissingModuleSource]
from datetime import datetime
from pathlib import Path
import sys
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side # pyright: ignore[reportMissingModuleSource]
from openpyxl.utils import get_column_letter # pyright: ignore[reportMissingModuleSource]

print("="*60)
print("🚀 GCP AUDIT EVIDENCE DEMO - EDUCATIONAL VERSION")
print("="*60)

# Set up paths
BASE_DIR = Path(__file__).parent.parent
MOCK_DATA = BASE_DIR / "mock_data"
OUTPUT_DIR = BASE_DIR / "output"
FRAMEWORK_MAP = BASE_DIR / "framework_map.csv"

# Create output directory
OUTPUT_DIR.mkdir(exist_ok=True)

# Define color schemes
SEVERITY_COLORS = {
    'CRITICAL': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),  # Red
    'HIGH': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),      # Orange
    'MEDIUM': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),    # Yellow
    'LOW': PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'),       # Light Green
    'INFO': PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')       # Light Blue
}

TAB_COLORS = {
    'Executive_Summary': 'FFC000',      # Gold
    'GCP_Findings': '5B9BD5',          # Blue
    'SIEM_Events': 'ED7D31',           # Orange
    'ScubaGoggles': 'FF6B35',          # Red-Orange
    'Framework_Mappings': '70AD47',     # Green
    'Access_Controls': '9A6DD7',        # Purple
    'Data_Security': '4472C4',          # Dark Blue
    'Network_Security': '2E75B5',       # Medium Blue
    'Audit_Logging': 'FFC000',          # Gold
    'ScubaGoggles_Assessment': 'FF6B35', # Red-Orange
    'SIEM_Incidents': 'ED7D31',         # Orange
    'Vulnerability_Management': 'E15759', # Red
    'System_Maintenance': '76B7B2',     # Teal
    'Backup_Recovery': '59A14F'         # Green
}

def load_findings():
    """Load mock GCP Security Command Center findings"""
    findings_file = MOCK_DATA / "findings.json"
    if not findings_file.exists():
        findings_file.parent.mkdir(exist_ok=True)
        mock_findings = {
            "findings": [
                {
                    "rule_id": "DEMO_001",
                    "severity": "HIGH",
                    "resource_name": "compute-instance-001",
                    "project_id": "demo-project",
                    "finding_class": "MISCONFIGURATION",
                    "description": "Instance with public IP and no firewall rules",
                    "event_time": "2025-01-15T10:00:00Z"
                },
                {
                    "rule_id": "DEMO_002",
                    "severity": "MEDIUM",
                    "resource_name": "storage-bucket-public",
                    "project_id": "demo-project",
                    "finding_class": "PUBLIC_ACCESS",
                    "description": "Storage bucket with public access enabled",
                    "event_time": "2025-01-15T11:00:00Z"
                }
            ]
        }
        with open(findings_file, 'w') as f:
            json.dump(mock_findings, f, indent=2)
    
    with open(findings_file) as f:
        data = json.load(f)
    return pd.DataFrame(data.get('findings', []))

def load_siem_events():
    """Load mock SIEM events"""
    siem_file = MOCK_DATA / "siem_events.json"
    if not siem_file.exists():
        mock_siem = {
            "siem_events": [
                {
                    "rule_id": "AUTH_BRUTE_FORCE",
                    "severity": "CRITICAL",
                    "event_type": "BRUTE_FORCE_ATTEMPT",
                    "source_ip": "192.168.1.100",
                    "description": "Multiple failed login attempts detected",
                    "event_time": "2025-01-15T12:00:00Z"
                }
            ]
        }
        with open(siem_file, 'w') as f:
            json.dump(mock_siem, f, indent=2)
    
    with open(siem_file) as f:
        data = json.load(f)
    return pd.DataFrame(data.get('siem_events', []))

def load_scuba_findings():
    """Load and transform CISA ScubaGoggles findings"""
    scuba_file = MOCK_DATA / "scuba_export.json"
    if not scuba_file.exists():
        mock_scuba = {
            "resources": [
                {
                    "name": "//storage.googleapis.com/public-bucket",
                    "type": "storage.googleapis.com/Bucket",
                    "project_id": "demo-project",
                    "iam": {
                        "bindings": [
                            {
                                "role": "roles/storage.objectViewer",
                                "members": ["allUsers"]
                            }
                        ]
                    }
                }
            ]
        }
        with open(scuba_file, 'w') as f:
            json.dump(mock_scuba, f, indent=2)
    
    with open(scuba_file) as f:
        data = json.load(f)
    
    # Transform ScubaGoggles data into findings format
    findings = []
    for resource in data.get('resources', []):
        if 'iam' in resource and 'bindings' in resource['iam']:
            for binding in resource['iam']['bindings']:
                if 'allUsers' in binding.get('members', []):
                    findings.append({
                        'rule_id': 'SCUBA_PUBLIC_ACCESS',
                        'severity': 'HIGH',
                        'resource_name': resource.get('name', 'Unknown'),
                        'project_id': resource.get('project_id', 'Unknown'),
                        'finding_class': 'MISCONFIGURATION',
                        'description': f"Public access detected via ScubaGoggles: {binding.get('role')}",
                        'event_time': datetime.now().isoformat()
                    })
    
    return pd.DataFrame(findings)

def load_framework_mappings():
    """Load compliance framework mappings"""
    if not FRAMEWORK_MAP.exists():
        print(f"❌ Framework mappings not found at {FRAMEWORK_MAP}")
        return pd.DataFrame()
    
    return pd.read_csv(FRAMEWORK_MAP)

def apply_formatting(writer, sheet_name, df, apply_severity_colors=False):
    """Apply professional formatting to Excel sheets"""
    worksheet = writer.sheets[sheet_name]
    
    # Header formatting
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Apply header formatting
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Apply severity coloring if requested
    if apply_severity_colors and 'severity' in df.columns:
        severity_col_idx = df.columns.get_loc('severity') + 1
        for row in range(2, len(df) + 2):
            severity_cell = worksheet.cell(row=row, column=severity_col_idx)
            severity_value = severity_cell.value
            if severity_value in SEVERITY_COLORS:
                severity_cell.fill = SEVERITY_COLORS[severity_value]
                severity_cell.font = Font(bold=True)
    
    # Auto-fit columns
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value or '')) for cell in column_cells)
        column_letter = column_cells[0].column_letter
        worksheet.column_dimensions[column_letter].width = min(length + 2, 50)
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
        for cell in row:
            cell.border = thin_border
    
    # Set tab color
    if sheet_name in TAB_COLORS:
        worksheet.sheet_properties.tabColor = TAB_COLORS[sheet_name]

def generate_excel_report(findings_df, siem_df, scuba_df, mappings_df):
    """Generate the audit-ready Excel report with professional formatting"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = OUTPUT_DIR / f"audit_report_{timestamp}.xlsx"
    
    # Merge with framework mappings
    findings_merged = findings_df.merge(mappings_df, on='rule_id', how='left')
    siem_merged = siem_df.merge(mappings_df, on='rule_id', how='left')
    scuba_merged = scuba_df.merge(mappings_df, on='rule_id', how='left')
    
    # Create Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Executive Summary
        summary_data = {
            'Data Source': ['GCP Findings', 'SIEM Events', 'ScubaGoggles', 'Total'],
            'Total Count': [
                len(findings_df), 
                len(siem_df), 
                len(scuba_df),
                len(findings_df) + len(siem_df) + len(scuba_df)
            ],
            'Critical': [
                len(findings_df[findings_df['severity'] == 'CRITICAL']),
                len(siem_df[siem_df['severity'] == 'CRITICAL']),
                len(scuba_df[scuba_df['severity'] == 'CRITICAL']),
                len(findings_df[findings_df['severity'] == 'CRITICAL']) + 
                len(siem_df[siem_df['severity'] == 'CRITICAL']) + 
                len(scuba_df[scuba_df['severity'] == 'CRITICAL'])
            ],
            'High': [
                len(findings_df[findings_df['severity'] == 'HIGH']),
                len(siem_df[siem_df['severity'] == 'HIGH']),
                len(scuba_df[scuba_df['severity'] == 'HIGH']),
                len(findings_df[findings_df['severity'] == 'HIGH']) + 
                len(siem_df[siem_df['severity'] == 'HIGH']) + 
                len(scuba_df[scuba_df['severity'] == 'HIGH'])
            ],
            'Medium': [
                len(findings_df[findings_df['severity'] == 'MEDIUM']),
                len(siem_df[siem_df['severity'] == 'MEDIUM']),
                len(scuba_df[scuba_df['severity'] == 'MEDIUM']),
                len(findings_df[findings_df['severity'] == 'MEDIUM']) + 
                len(siem_df[siem_df['severity'] == 'MEDIUM']) + 
                len(scuba_df[scuba_df['severity'] == 'MEDIUM'])
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Executive_Summary', index=False)
        apply_formatting(writer, 'Executive_Summary', summary_df, False)
        
        # Detailed findings sheets with formatting
        if not findings_merged.empty:
            findings_merged.to_excel(writer, sheet_name='GCP_Findings', index=False)
            apply_formatting(writer, 'GCP_Findings', findings_merged, True)
        
        if not siem_merged.empty:
            siem_merged.to_excel(writer, sheet_name='SIEM_Events', index=False)
            apply_formatting(writer, 'SIEM_Events', siem_merged, True)
        
        if not scuba_merged.empty:
            scuba_merged.to_excel(writer, sheet_name='ScubaGoggles', index=False)
            apply_formatting(writer, 'ScubaGoggles', scuba_merged, True)
        
        mappings_df.to_excel(writer, sheet_name='Framework_Mappings', index=False)
        apply_formatting(writer, 'Framework_Mappings', mappings_df, False)
        
        # Group findings by excel_tab from framework mappings
        all_findings = pd.concat([findings_merged, siem_merged, scuba_merged], ignore_index=True)
        
        # Create sheets based on excel_tab categories
        if 'excel_tab' in all_findings.columns:
            for tab in all_findings['excel_tab'].dropna().unique():
                tab_data = all_findings[all_findings['excel_tab'] == tab]
                if not tab_data.empty and tab not in ['GCP_Findings', 'SIEM_Events', 'ScubaGoggles']:
                    # Clean sheet name (Excel has 31 char limit)
                    clean_tab = tab[:31]
                    tab_data.to_excel(writer, sheet_name=clean_tab, index=False)
                    apply_formatting(writer, clean_tab, tab_data, True)
    
    return output_file

def main():
    """Main execution function"""
    try:
        print("\n📊 Loading data sources...")
        print("  ✓ Loading GCP findings...")
        findings = load_findings()
        print(f"    Found {len(findings)} findings")
        
        print("  ✓ Loading SIEM events...")
        siem = load_siem_events()
        print(f"    Found {len(siem)} events")
        
        print("  ✓ Loading ScubaGoggles assessments...")
        scuba = load_scuba_findings()
        print(f"    Found {len(scuba)} assessments")
        
        print("\n🔗 Loading framework mappings...")
        mappings = load_framework_mappings()
        print(f"    Loaded {len(mappings)} control mappings")
        
        print("\n📝 Generating Excel report with formatting...")
        output_file = generate_excel_report(findings, siem, scuba, mappings)
        
        print("\n✨ SUCCESS! Report generated:")
        print(f"    📁 {output_file}")
        print("\n📊 Report features:")
        print("  • Color-coded severity levels (Critical=Red, High=Orange, Medium=Yellow)")
        print("  • Colored worksheet tabs by category")
        print("  • Professional formatting with borders and headers")
        print("  • Auto-organized by compliance categories")
        print("\n" + "="*60)
        print("Demo complete! Open the Excel file to see the formatted results.")
        print("\nThis educational demo shows the core concept.")
        print("Production version includes:")
        print("  • Real-time GCP API integration")
        print("  • 1000+ control mappings")
        print("  • Automated scheduling")
        print("  • Enterprise authentication")
        print("\nContact: info@eagledefensesys.com")
        print("="*60)
        
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0

if __name__ == "__main__":
    sys.exit(main())